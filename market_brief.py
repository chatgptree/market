"""
market_brief.py  —  Daily AI-powered market brief → Excel + HTML
Requirements:  pip install yfinance openpyxl anthropic
"""

import os, sys, json, re, time, datetime
import yfinance as yf
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────
# CONFIGURATION  ← edit this section
# ─────────────────────────────────────────────

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "YOUR_API_KEY_HERE")
OUTPUT_EXCEL = "Market_Brief.xlsx"
OUTPUT_HTML  = "index.html"   # GitHub Pages serves index.html automatically

PORTFOLIO = {
    "FCX":    {"units": None, "avg_aud": None,  "note": "Copper / AI power demand"},
    "CEG":    {"units": None, "avg_aud": None,  "note": "Nuclear / AI power"},
    "VST":    {"units": None, "avg_aud": None,  "note": "Nuclear / AI power"},
    "CSL.AX": {"units": 22,   "avg_aud": 149.0, "note": "Pharma - binary Aug 18 result"},
    "CI":     {"units": None, "avg_aud": None,  "note": "Defensive / uncorrelated"},
    "WGX.AX": {"units": 354,  "avg_aud": 4.93,  "note": "Gold - unhedged, FY result Sep 2"},
}

WATCHLIST = ["SLX.AX", "IREN", "MARA", "MRVL", "MU", "ALAB", "BIDU", "GOOG"]

MACRO = {
    "Gold":      "GC=F",
    "Copper":    "HG=F",
    "Oil (WTI)": "CL=F",
    "10Y Yield": "^TNX",
    "VIX":       "^VIX",
    "AUD/USD":   "AUDUSD=X",
    "S&P 500":   "^GSPC",
    "Nasdaq":    "^IXIC",
    "PHLX Semi": "^SOX",
    "ASX 200":   "^AXJO",
}

# ─────────────────────────────────────────────
# COLOUR PALETTE
# ─────────────────────────────────────────────

C = {
    "bg_dark":    "0D1117",
    "bg_mid":     "161B22",
    "bg_card":    "1C2128",
    "accent":     "00D4FF",
    "green":      "39D353",
    "red":        "F85149",
    "gold":       "E3B341",
    "text_white": "F0F6FC",
    "text_dim":   "8B949E",
}

def hf(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def mkborder():
    s = Side(style="thin", color="30363D")
    return Border(left=s, right=s, top=s, bottom=s)

def sh(cell, bg=None, fg=None, size=9, bold=True):
    cell.font = Font(name="Consolas", bold=bold, color=fg or C["text_white"], size=size)
    cell.fill = hf(bg or C["bg_dark"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = mkborder()

def sd(cell, bg=None, fg=None, bold=False, align="right"):
    cell.font = Font(name="Consolas", color=fg or C["text_white"], size=9, bold=bold)
    cell.fill = hf(bg or C["bg_card"])
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = mkborder()

# ─────────────────────────────────────────────
# DATA FETCHING
# ─────────────────────────────────────────────

def fetch_price_data(tickers):
    results = {}
    for ticker in tickers:
        for attempt in range(3):
            try:
                t   = yf.Ticker(ticker)
                fi  = t.fast_info
                inf = t.info
                price = fi.last_price
                prev  = fi.previous_close
                if price is None:
                    raise ValueError("price is None")
                chg  = ((price - prev) / prev * 100) if prev else 0
                hi52 = inf.get("fiftyTwoWeekHigh") or inf.get("52WeekHigh")
                lo52 = inf.get("fiftyTwoWeekLow")  or inf.get("52WeekLow")
                results[ticker] = {
                    "price":      round(price, 3),
                    "prev_close": round(prev, 3) if prev else "N/A",
                    "change_pct": round(chg, 2),
                    "day_high":   round(fi.day_high, 3) if fi.day_high else "N/A",
                    "day_low":    round(fi.day_low,  3) if fi.day_low  else "N/A",
                    "52w_high":   round(hi52, 3) if hi52 else "N/A",
                    "52w_low":    round(lo52, 3) if lo52 else "N/A",
                }
                break
            except Exception as e:
                if attempt < 2:
                    time.sleep(2)
                else:
                    print(f"      WARNING: Could not fetch {ticker}: {e}")
                    results[ticker] = {"price": "N/A", "change_pct": 0}
    return results

def build_context(portfolio_data, macro_data, watchlist_data, audusd):
    today = datetime.date.today().strftime("%A %d %B %Y")
    lines = [f"DATE: {today}", "", "=== MACRO ==="]
    for name, data in macro_data.items():
        chg = data.get("change_pct", 0)
        lines.append(f"{name:15} {str(data.get('price','N/A')):>10}  {'UP' if chg>0 else 'DOWN'} {abs(chg):.2f}%")
    lines += ["", "=== PORTFOLIO ==="]
    for ticker, cfg in PORTFOLIO.items():
        d     = portfolio_data.get(ticker, {})
        price = d.get("price", "N/A")
        chg   = d.get("change_pct", 0)
        avg   = cfg.get("avg_aud")
        units = cfg.get("units")
        pnl   = f"P&L: {(price-avg)/avg*100:+.1f}%" if avg and units and isinstance(price,(int,float)) else "P&L: N/A"
        lines.append(f"{ticker:10} Price:{str(price):>10}  Today:{chg:+.2f}%  {pnl}  | {cfg.get('note','')}")
    lines += ["", "=== WATCHLIST ==="]
    for ticker in WATCHLIST:
        d = watchlist_data.get(ticker, {})
        lines.append(f"{ticker:10} Price:{str(d.get('price','N/A')):>10}  Today:{d.get('change_pct',0):+.2f}%")
    aud_str = f"{audusd:.4f}" if isinstance(audusd,(int,float)) else str(audusd)
    lines += ["", f"AUD/USD: {aud_str}"]
    return "\n".join(lines)

# ─────────────────────────────────────────────
# CLAUDE ANALYSIS
# ─────────────────────────────────────────────

SYSTEM_PROMPT = """You are a blunt macro-aware investment analyst for a sophisticated retail investor (Giacky) in Melbourne, Australia. AUD base. Trading 212 platform (0.80% FX round-trip). No disclaimers. Flag rationalisations.

PORTFOLIO: FCX (copper/AI power), CEG (nuclear), VST (nuclear), CSL.AX (22 shares avg A$149 - underwater, binary catalyst Aug 18), WGX.AX (354 units avg A$4.93 - unhedged gold, FY result Sep 2), NDQ (Betashares ETF, weekly DCA).

MACRO FRAMEWORK: Iran/Hormuz disruption central thesis. Stagflation regime: sticky CPI 3.8%, Fed on hold, RBA hiking. AI bubble repricing risk Q4 2026 to Q2 2027 (25-40% on AI names). Portfolio 72% correlated to AI demand. Gold supercycle. Structural copper deficit. Nuclear renaissance.

WATCHLIST: SLX.AX (uranium enrichment, tranched entry), IREN (AI infrastructure pivot), MARA (watchlist only), MRVL (wait for Aug earnings), MU, ALAB, BIDU, GOOG.

CRITICAL RULES:
- Respond ONLY with a single valid JSON object. No text before or after.
- No markdown fences. No preamble.
- Do NOT use apostrophes or contractions in string values. Write "does not" not "doesn't". Write "it is" not "it's".
- Do NOT use single quotes anywhere inside the JSON.

JSON STRUCTURE:
{
  "regime": "2-3 sentence macro regime snapshot",
  "sector_rotation": "1-2 sentences on where money is moving",
  "portfolio_alerts": [
    {"ticker": "XXX", "signal": "HOLD", "reason": "one sentence without apostrophes"}
  ],
  "top_ideas": [
    {"rank": 1, "ticker": "XXX", "thesis": "2 sentences", "entry": "price or range", "sizing": "percent or AUD amount", "invalidation": "level or event"}
  ],
  "risks": "2 sentences on biggest near-term risk",
  "watchlist_notes": "1-2 sentences on watchlist names worth flagging"
}"""

def get_claude_analysis(context):
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    msg = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1500,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": f"Live market data:\n\n{context}\n\nProvide analysis now."}]
    )
    raw = msg.content[0].text.strip()
    raw = raw.replace("```json","").replace("```","").strip()
    start = raw.find("{")
    end   = raw.rfind("}") + 1
    if start == -1 or end == 0:
        raise ValueError(f"No JSON found: {raw[:200]}")
    raw = raw[start:end]
    # attempt 1: parse directly
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        pass
    # attempt 2: strip control characters — newlines inside strings cause comma errors
    cleaned = re.sub(r'[\x00-\x1f\x7f]', ' ', raw)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass
    # attempt 3: normalise fancy quotes
    cleaned = cleaned.replace('\u2018',' ').replace('\u2019',' ')
    cleaned = cleaned.replace('\u201c','"').replace('\u201d','"')
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass
    # attempt 4: strip apostrophes inside JSON string values
    def remove_apostrophes(m):
        return m.group(0).replace("'", " ")
    cleaned = re.sub(r'"[^"\\n]*"', remove_apostrophes, cleaned)
    return json.loads(cleaned)

# ─────────────────────────────────────────────
# HTML GENERATOR
# ─────────────────────────────────────────────

def chg_color(chg):
    if isinstance(chg, (int, float)):
        return "#39D353" if chg > 0 else ("#F85149" if chg < 0 else "#8B949E")
    return "#8B949E"

def chg_arrow(chg):
    if isinstance(chg, (int, float)):
        return "▲" if chg > 0 else ("▼" if chg < 0 else "—")
    return "—"

def fmt_chg(chg):
    if isinstance(chg, (int, float)):
        return f"{chg:+.2f}%"
    return "N/A"

def signal_color(signal):
    return {"ADD":"#39D353","TRIM":"#E3B341","HOLD":"#8B949E","WATCH":"#00D4FF"}.get(signal.upper(),"#F0F6FC")

def build_html(analysis, macro_data, portfolio_data, watchlist_data, audusd, today_str):
    aud_str = f"{audusd:.4f}" if isinstance(audusd,(int,float)) else str(audusd)

    # ── macro rows ──
    macro_rows = ""
    for name, data in macro_data.items():
        price = data.get("price","N/A")
        chg   = data.get("change_pct", 0)
        hi52  = data.get("52w_high","N/A")
        macro_rows += f"""
        <tr>
          <td style="color:#F0F6FC;font-weight:600">{name}</td>
          <td style="color:#F0F6FC;text-align:right">{price}</td>
          <td style="color:{chg_color(chg)};text-align:right">{chg_arrow(chg)} {fmt_chg(chg)}</td>
          <td style="color:#8B949E;text-align:right">{hi52}</td>
        </tr>"""

    # ── portfolio rows ──
    port_rows = ""
    audusd_rate = audusd if isinstance(audusd,(int,float)) else 0.65
    total_inv = 0
    total_val = 0
    for ticker, cfg in PORTFOLIO.items():
        d     = portfolio_data.get(ticker, {})
        price = d.get("price","N/A")
        chg   = d.get("change_pct", 0)
        units = cfg.get("units")
        avg   = cfg.get("avg_aud")
        note  = cfg.get("note","")
        if price and units and avg and isinstance(price,(int,float)):
            price_aud = price if ticker.endswith(".AX") else price / audusd_rate
            mkt_val   = round(price_aud * units, 2)
            invested  = round(avg * units, 2)
            pnl_pct   = round((price_aud - avg) / avg * 100, 2)
            total_inv += invested
            total_val += mkt_val
            pnl_str   = f"{pnl_pct:+.2f}%"
            val_str   = f"A${mkt_val:,.0f}"
            pnl_col   = chg_color(pnl_pct)
        else:
            pnl_str = "N/A"
            val_str = "N/A"
            pnl_col = "#8B949E"
        port_rows += f"""
        <tr>
          <td style="color:#00D4FF;font-weight:700">{ticker}</td>
          <td style="color:#8B949E;text-align:right">{units if units else '—'}</td>
          <td style="color:#8B949E;text-align:right">{'A$'+str(avg) if avg else '—'}</td>
          <td style="color:#F0F6FC;text-align:right">{price}</td>
          <td style="color:{chg_color(chg)};text-align:right">{fmt_chg(chg)}</td>
          <td style="color:{pnl_col};text-align:right;font-weight:600">{pnl_str}</td>
          <td style="color:#F0F6FC;text-align:right">{val_str}</td>
          <td style="color:#8B949E;font-size:11px">{note}</td>
        </tr>"""

    tot_pnl = round((total_val - total_inv) / total_inv * 100, 2) if total_inv else 0
    tot_col = "#39D353" if tot_pnl > 0 else "#F85149"
    port_rows += f"""
        <tr style="border-top:1px solid #30363D">
          <td colspan="6" style="color:#E3B341;font-weight:700;padding-top:8px">TOTAL PORTFOLIO</td>
          <td style="color:{tot_col};text-align:right;font-weight:700">A${total_val:,.0f}<br><span style="font-size:11px">({tot_pnl:+.1f}%)</span></td>
          <td></td>
        </tr>"""

    # ── watchlist rows ──
    watch_rows = ""
    for ticker in WATCHLIST:
        d     = watchlist_data.get(ticker, {})
        price = d.get("price","N/A")
        chg   = d.get("change_pct", 0)
        hi52  = d.get("52w_high","N/A")
        lo52  = d.get("52w_low","N/A")
        vs_high = round((price-hi52)/hi52*100,1) if isinstance(price,(int,float)) and isinstance(hi52,(int,float)) and hi52 else "N/A"
        vs_col  = "#39D353" if isinstance(vs_high,float) and vs_high>-10 else ("#E3B341" if isinstance(vs_high,float) and vs_high>-25 else "#F85149")
        watch_rows += f"""
        <tr>
          <td style="color:#00D4FF;font-weight:700">{ticker}</td>
          <td style="color:#F0F6FC;text-align:right">{price}</td>
          <td style="color:{chg_color(chg)};text-align:right">{chg_arrow(chg)} {fmt_chg(chg)}</td>
          <td style="color:#8B949E;text-align:right">{hi52}</td>
          <td style="color:#8B949E;text-align:right">{lo52}</td>
          <td style="color:{vs_col};text-align:right;font-weight:600">{vs_high if vs_high=='N/A' else str(vs_high)+'%'}</td>
        </tr>"""

    # ── alerts ──
    alerts_html = ""
    for alert in analysis.get("portfolio_alerts", []):
        sig = alert.get("signal","")
        alerts_html += f"""
        <div style="display:flex;align-items:flex-start;gap:12px;padding:10px 0;border-bottom:1px solid #21262d">
          <span style="color:#00D4FF;font-weight:700;min-width:80px">{alert.get('ticker','')}</span>
          <span style="color:{signal_color(sig)};font-weight:700;min-width:60px">[{sig}]</span>
          <span style="color:#F0F6FC">{alert.get('reason','')}</span>
        </div>"""

    # ── top ideas ──
    ideas_html = ""
    rank_colors = ["#E3B341","#F0F6FC","#8B949E"]
    for idea in analysis.get("top_ideas",[]):
        rank = idea.get("rank",1)
        rc   = rank_colors[min(rank-1,2)]
        ideas_html += f"""
        <div style="background:#1C2128;border:1px solid #30363D;border-radius:8px;padding:16px;margin-bottom:12px">
          <div style="display:flex;align-items:center;gap:12px;margin-bottom:10px">
            <span style="color:{rc};font-size:20px;font-weight:900">#{rank}</span>
            <span style="color:#00D4FF;font-size:16px;font-weight:700">{idea.get('ticker','')}</span>
          </div>
          <p style="color:#F0F6FC;margin:0 0 10px">{idea.get('thesis','')}</p>
          <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px">
            <div style="background:#0D1117;border-radius:4px;padding:8px">
              <div style="color:#8B949E;font-size:10px;text-transform:uppercase;margin-bottom:2px">Entry</div>
              <div style="color:#39D353;font-weight:600">{idea.get('entry','')}</div>
            </div>
            <div style="background:#0D1117;border-radius:4px;padding:8px">
              <div style="color:#8B949E;font-size:10px;text-transform:uppercase;margin-bottom:2px">Sizing</div>
              <div style="color:#E3B341;font-weight:600">{idea.get('sizing','')}</div>
            </div>
            <div style="background:#0D1117;border-radius:4px;padding:8px">
              <div style="color:#8B949E;font-size:10px;text-transform:uppercase;margin-bottom:2px">Invalidation</div>
              <div style="color:#F85149;font-weight:600">{idea.get('invalidation','')}</div>
            </div>
          </div>
        </div>"""

    table_style = "width:100%;border-collapse:collapse;font-size:13px"
    th_style    = "background:#161B22;color:#8B949E;font-size:10px;text-transform:uppercase;letter-spacing:.05em;padding:8px 12px;text-align:left;border-bottom:1px solid #30363D"
    td_css      = "padding:8px 12px;border-bottom:1px solid #21262d"

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Market Brief // {today_str}</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{background:#0D1117;color:#F0F6FC;font-family:'JetBrains Mono','Fira Code','Consolas',monospace;font-size:13px;line-height:1.6}}
  a{{color:#00D4FF;text-decoration:none}}
  h2{{color:#E3B341;font-size:11px;text-transform:uppercase;letter-spacing:.1em;margin-bottom:16px;padding-bottom:6px;border-bottom:1px solid #30363D}}
  .container{{max-width:1100px;margin:0 auto;padding:24px 16px}}
  .header{{border-bottom:1px solid #30363D;padding-bottom:20px;margin-bottom:28px}}
  .header-title{{font-size:clamp(18px,4vw,28px);font-weight:900;color:#00D4FF;letter-spacing:.05em}}
  .header-date{{color:#8B949E;font-size:12px;margin-top:4px}}
  .regime-box{{background:#0f1f30;border-left:3px solid #00D4FF;border-radius:0 6px 6px 0;padding:14px 18px;margin-bottom:28px;color:#F0F6FC;font-style:italic;line-height:1.7}}
  .grid-2{{display:grid;grid-template-columns:1fr 1fr;gap:24px;margin-bottom:28px}}
  .grid-3{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:16px;margin-bottom:28px}}
  @media(max-width:700px){{.grid-2,.grid-3{{grid-template-columns:1fr}}}}
  .card{{background:#161B22;border:1px solid #30363D;border-radius:8px;padding:20px}}
  table td{{padding:8px 12px;border-bottom:1px solid #21262d}}
  table tr:last-child td{{border-bottom:none}}
  table tr:hover td{{background:#1C2128}}
  .stat-card{{background:#1C2128;border:1px solid #30363D;border-radius:8px;padding:16px;text-align:center}}
  .stat-label{{color:#8B949E;font-size:10px;text-transform:uppercase;letter-spacing:.05em;margin-bottom:4px}}
  .stat-value{{font-size:20px;font-weight:900}}
  .footer{{border-top:1px solid #30363D;padding-top:16px;margin-top:32px;color:#8B949E;font-size:11px;text-align:center}}
  .tag{{display:inline-block;padding:2px 8px;border-radius:3px;font-size:10px;font-weight:700;text-transform:uppercase}}
</style>
</head>
<body>
<div class="container">

  <!-- HEADER -->
  <div class="header">
    <div class="header-title">MARKET BRIEF</div>
    <div class="header-date">{today_str} &nbsp;·&nbsp; Powered by Claude Sonnet 4.6 &nbsp;·&nbsp; Live via yFinance &nbsp;·&nbsp; AUD base</div>
  </div>

  <!-- REGIME -->
  <div class="regime-box">
    <span style="color:#00D4FF;font-weight:700;font-style:normal">REGIME &nbsp;</span>{analysis.get("regime","N/A")}
  </div>

  <!-- KEY STATS -->
  <div class="grid-3" style="margin-bottom:28px">
    <div class="stat-card">
      <div class="stat-label">AUD / USD</div>
      <div class="stat-value" style="color:#E3B341">{aud_str}</div>
    </div>
    <div class="stat-card">
      <div class="stat-label">VIX</div>
      <div class="stat-value" style="color:#F0F6FC">{macro_data.get('VIX',{{}}).get('price','N/A')}</div>
    </div>
    <div class="stat-card">
      <div class="stat-label">10Y Yield</div>
      <div class="stat-value" style="color:#F0F6FC">{macro_data.get('10Y Yield',{{}}).get('price','N/A')}</div>
    </div>
  </div>

  <!-- MACRO + PORTFOLIO SIDE BY SIDE -->
  <div class="grid-2">
    <div class="card">
      <h2>Macro</h2>
      <table style="{table_style}">
        <thead><tr>
          <th style="{th_style}">Instrument</th>
          <th style="{th_style};text-align:right">Price</th>
          <th style="{th_style};text-align:right">Change</th>
          <th style="{th_style};text-align:right">52W High</th>
        </tr></thead>
        <tbody>{macro_rows}</tbody>
      </table>
    </div>
    <div class="card">
      <h2>Portfolio Alerts</h2>
      {alerts_html if alerts_html else '<p style="color:#8B949E">No alerts today.</p>'}
      <div style="margin-top:20px">
        <h2>Sector Rotation</h2>
        <p style="color:#F0F6FC">{analysis.get('sector_rotation','N/A')}</p>
      </div>
      <div style="margin-top:20px">
        <h2>Risks</h2>
        <p style="color:#F85149">{analysis.get('risks','N/A')}</p>
      </div>
    </div>
  </div>

  <!-- TOP IDEAS -->
  <div class="card" style="margin-bottom:24px">
    <h2>Top 3 Actionable Ideas</h2>
    {ideas_html if ideas_html else '<p style="color:#8B949E">No ideas generated.</p>'}
  </div>

  <!-- PORTFOLIO P&L -->
  <div class="card" style="margin-bottom:24px">
    <h2>Portfolio Positions &amp; P&amp;L</h2>
    <div style="overflow-x:auto">
      <table style="{table_style}">
        <thead><tr>
          <th style="{th_style}">Ticker</th>
          <th style="{th_style};text-align:right">Units</th>
          <th style="{th_style};text-align:right">Avg Cost</th>
          <th style="{th_style};text-align:right">Price</th>
          <th style="{th_style};text-align:right">Today</th>
          <th style="{th_style};text-align:right">P&amp;L</th>
          <th style="{th_style};text-align:right">Mkt Value</th>
          <th style="{th_style}">Notes</th>
        </tr></thead>
        <tbody>{port_rows}</tbody>
      </table>
    </div>
  </div>

  <!-- WATCHLIST -->
  <div class="card" style="margin-bottom:24px">
    <h2>Watchlist</h2>
    <div style="overflow-x:auto">
      <table style="{table_style}">
        <thead><tr>
          <th style="{th_style}">Ticker</th>
          <th style="{th_style};text-align:right">Price</th>
          <th style="{th_style};text-align:right">Today</th>
          <th style="{th_style};text-align:right">52W High</th>
          <th style="{th_style};text-align:right">52W Low</th>
          <th style="{th_style};text-align:right">vs 52W High</th>
        </tr></thead>
        <tbody>{watch_rows}</tbody>
      </table>
    </div>
    <div style="margin-top:16px;padding-top:16px;border-top:1px solid #30363D;color:#F0F6FC">
      <span style="color:#8B949E;font-size:11px;text-transform:uppercase">Watchlist Notes &nbsp;</span>{analysis.get('watchlist_notes','N/A')}
    </div>
  </div>

  <div class="footer">
    Generated {today_str} &nbsp;·&nbsp; Claude Sonnet 4.6 &nbsp;·&nbsp; yFinance &nbsp;·&nbsp; For informational purposes only
  </div>

</div>
</body>
</html>"""
    return html

# ─────────────────────────────────────────────
# EXCEL BUILDER
# ─────────────────────────────────────────────

def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def title_row(ws, row, text, end_col, bg=None, fg=None, size=11):
    ws.merge_cells(f"A{row}:{end_col}{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font  = Font(name="Consolas", bold=True, color=fg or C["accent"], size=size)
    c.fill  = hf(bg or C["bg_dark"])
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22

def fill_bg(ws, rows, cols):
    for row in ws.iter_rows(min_row=1, max_row=rows, min_col=1, max_col=cols):
        for cell in row:
            cell.fill = hf(C["bg_dark"])

def build_dashboard(wb, analysis, macro_data, portfolio_data, audusd, today_str):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"
    fill_bg(ws, 60, 20)
    set_widths(ws, {"A":2,"B":18,"C":14,"D":12,"E":12,"F":2,"G":18,"H":14,"I":12,"J":12})

    ws.merge_cells("B1:J1")
    c = ws["B1"]
    c.value = f"  MARKET BRIEF  //  {today_str}"
    c.font  = Font(name="Consolas", bold=True, color=C["accent"], size=14)
    c.fill  = hf(C["bg_dark"])
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("B2:J2")
    c = ws["B2"]
    c.value = "  Powered by Claude Sonnet 4.6  |  Live via yFinance  |  AUD base"
    c.font  = Font(name="Consolas", color=C["text_dim"], size=9)
    c.fill  = hf(C["bg_dark"])
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    ws.merge_cells("B4:J4")
    c = ws["B4"]
    c.value = "  REGIME: " + analysis.get("regime","N/A")
    c.font  = Font(name="Consolas", color=C["text_white"], size=9, italic=True)
    c.fill  = hf("1A2332")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 44

    title_row(ws, 6, "  MACRO", "E", size=9)
    for i, h in enumerate(["INSTRUMENT","PRICE","CHG %","52W HIGH"]):
        c = ws.cell(row=7, column=2+i, value=h)
        sh(c, bg=C["bg_mid"], fg=C["text_dim"], size=8)
    ws.row_dimensions[7].height = 18

    for r, (name, data) in enumerate(macro_data.items(), start=8):
        chg    = data.get("change_pct", 0)
        fg_chg = C["green"] if chg>0 else (C["red"] if chg<0 else C["text_dim"])
        row_bg = C["bg_card"] if r%2==0 else C["bg_mid"]
        for i, (v, fc) in enumerate(zip(
            [name, data.get("price","N/A"), f"{chg:+.2f}%", data.get("52w_high","N/A")],
            [C["text_white"], C["text_white"], fg_chg, C["text_dim"]]
        )):
            c = ws.cell(row=r, column=2+i, value=v)
            sd(c, bg=row_bg, fg=fc, align="left" if i==0 else "right")
        ws.row_dimensions[r].height = 16

    title_row(ws, 6, "  PORTFOLIO", "J", bg=C["bg_dark"])
    for i, h in enumerate(["TICKER","PRICE","TODAY","AVG COST"]):
        c = ws.cell(row=7, column=7+i, value=h)
        sh(c, bg=C["bg_mid"], fg=C["text_dim"], size=8)

    for r, (ticker, cfg) in enumerate(PORTFOLIO.items(), start=8):
        d      = portfolio_data.get(ticker, {})
        price  = d.get("price","N/A")
        chg    = d.get("change_pct", 0)
        avg    = cfg.get("avg_aud")
        fg_chg = C["green"] if chg>0 else (C["red"] if chg<0 else C["text_dim"])
        row_bg = C["bg_card"] if r%2==0 else C["bg_mid"]
        for i, (v, fc) in enumerate(zip(
            [ticker, price, f"{chg:+.2f}%", avg if avg else "---"],
            [C["accent"], C["text_white"], fg_chg, C["text_dim"]]
        )):
            c = ws.cell(row=r, column=7+i, value=v)
            sd(c, bg=row_bg, fg=fc, align="left" if i==0 else "right")
        ws.row_dimensions[r].height = 16

    aud_row = max(8+len(macro_data), 8+len(PORTFOLIO)) + 1
    aud_str = f"{audusd:.4f}" if isinstance(audusd,(int,float)) else str(audusd)
    ws.merge_cells(f"B{aud_row}:J{aud_row}")
    c = ws[f"B{aud_row}"]
    c.value = f"  AUD/USD  {aud_str}   |   Sector rotation: {analysis.get('sector_rotation','')}"
    c.font  = Font(name="Consolas", color=C["gold"], size=9)
    c.fill  = hf(C["bg_mid"])
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[aud_row].height = 18

def build_brief(wb, analysis):
    ws = wb.create_sheet("Daily Brief")
    ws.sheet_view.showGridLines = False
    fill_bg(ws, 80, 14)
    set_widths(ws, {"A":2,"B":16,"C":60,"D":2})
    ws.row_dimensions[1].height = 30

    ws.merge_cells("B1:C1")
    c = ws["B1"]
    c.value = "  CLAUDE DAILY BRIEF"
    c.font  = Font(name="Consolas", bold=True, color=C["accent"], size=13)
    c.fill  = hf(C["bg_dark"])
    c.alignment = Alignment(horizontal="left", vertical="center")

    row = 3
    for label, content in [
        ("REGIME",          analysis.get("regime","")),
        ("SECTOR ROTATION", analysis.get("sector_rotation","")),
        ("RISKS",           analysis.get("risks","")),
        ("WATCHLIST NOTES", analysis.get("watchlist_notes","")),
    ]:
        ws.merge_cells(f"B{row}:C{row}")
        h = ws[f"B{row}"]
        h.value = f"  {label}"
        h.font  = Font(name="Consolas", bold=True, color=C["gold"], size=9)
        h.fill  = hf(C["bg_mid"])
        h.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1
        ws.merge_cells(f"B{row}:C{row}")
        d = ws[f"B{row}"]
        d.value = "  " + content
        d.font  = Font(name="Consolas", color=C["text_white"], size=9)
        d.fill  = hf(C["bg_card"])
        d.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 44
        row += 2

    ws.merge_cells(f"B{row}:C{row}")
    h = ws[f"B{row}"]
    h.value = "  PORTFOLIO ALERTS"
    h.font  = Font(name="Consolas", bold=True, color=C["gold"], size=9)
    h.fill  = hf(C["bg_mid"])
    h.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18
    row += 1

    sig_colors = {"ADD":C["green"],"TRIM":C["gold"],"HOLD":C["text_dim"],"WATCH":C["accent"]}
    for alert in analysis.get("portfolio_alerts",[]):
        sig = alert.get("signal","")
        c1  = ws.cell(row=row, column=2, value=alert.get("ticker",""))
        sd(c1, bg=C["bg_card"], fg=C["accent"], bold=True, align="left")
        c2  = ws.cell(row=row, column=3, value=f"[{sig}]  {alert.get('reason','')}")
        sd(c2, bg=C["bg_card"], fg=sig_colors.get(sig.upper(), C["text_white"]), align="left")
        ws.row_dimensions[row].height = 20
        row += 1

def build_ideas(wb, analysis):
    ws = wb.create_sheet("Top Ideas")
    ws.sheet_view.showGridLines = False
    fill_bg(ws, 60, 14)
    set_widths(ws, {"A":2,"B":12,"C":40,"D":18,"E":20,"F":28})

    ws.merge_cells("B1:F1")
    c = ws["B1"]
    c.value = "  TOP 3 ACTIONABLE IDEAS"
    c.font  = Font(name="Consolas", bold=True, color=C["accent"], size=13)
    c.fill  = hf(C["bg_dark"])
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    for i, h in enumerate(["TICKER","THESIS","ENTRY","SIZING","INVALIDATION"]):
        c = ws.cell(row=3, column=2+i, value=h)
        sh(c, bg=C["bg_mid"], fg=C["text_dim"], size=8)
    ws.row_dimensions[3].height = 18

    rank_colors = [C["gold"], C["text_white"], C["text_dim"]]
    for idea in analysis.get("top_ideas",[]):
        rank    = idea.get("rank",1)
        row_num = 3 + rank
        row_bg  = C["bg_card"] if rank%2==0 else "1A2332"
        fc      = rank_colors[min(rank-1,2)]
        for i, v in enumerate([idea.get("ticker",""), idea.get("thesis",""), idea.get("entry",""), idea.get("sizing",""), idea.get("invalidation","")]):
            c = ws.cell(row=row_num, column=2+i, value=v)
            sd(c, bg=row_bg, fg=fc if i==0 else C["text_white"], align="left")
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row_num].height = 44

def build_portfolio_sheet(wb, portfolio_data, audusd):
    ws = wb.create_sheet("Portfolio")
    ws.sheet_view.showGridLines = False
    fill_bg(ws, 40, 14)
    set_widths(ws, {"A":2,"B":14,"C":10,"D":12,"E":12,"F":12,"G":12,"H":14,"I":30})

    ws.merge_cells("B1:I1")
    c = ws["B1"]
    c.value = "  PORTFOLIO POSITIONS & P&L"
    c.font  = Font(name="Consolas", bold=True, color=C["accent"], size=13)
    c.fill  = hf(C["bg_dark"])
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    for i, h in enumerate(["TICKER","UNITS","AVG COST","CURR PRICE","TODAY %","P&L %","MKT VALUE","NOTES"]):
        c = ws.cell(row=3, column=2+i, value=h)
        sh(c, bg=C["bg_mid"], fg=C["text_dim"], size=8)
    ws.row_dimensions[3].height = 18

    audusd_rate = audusd if isinstance(audusd,(int,float)) else 0.65
    total_inv = 0
    total_val = 0

    for r, (ticker, cfg) in enumerate(PORTFOLIO.items(), start=4):
        d     = portfolio_data.get(ticker, {})
        price = d.get("price")
        chg   = d.get("change_pct", 0)
        units = cfg.get("units")
        avg   = cfg.get("avg_aud")
        note  = cfg.get("note","")

        if price and units and avg and isinstance(price,(int,float)):
            price_aud = price if ticker.endswith(".AX") else price / audusd_rate
            mkt_val   = round(price_aud * units, 2)
            invested  = round(avg * units, 2)
            pnl_pct   = round((price_aud - avg) / avg * 100, 2)
            total_inv += invested
            total_val += mkt_val
        else:
            price_aud = price
            mkt_val   = "N/A"
            pnl_pct   = "N/A"

        row_bg   = C["bg_card"] if r%2==0 else C["bg_mid"]
        fg_today = C["green"] if chg>0 else (C["red"] if chg<0 else C["text_dim"])
        fg_pnl   = C["green"] if isinstance(pnl_pct,float) and pnl_pct>0 else (C["red"] if isinstance(pnl_pct,float) and pnl_pct<0 else C["text_dim"])

        for i, (v, fc) in enumerate([
            (ticker,                                                         C["accent"]),
            (units or "---",                                                 C["text_white"]),
            (f"A${avg:.2f}" if avg else "---",                              C["text_dim"]),
            (f"{price_aud:.3f}" if isinstance(price_aud,float) else "N/A", C["text_white"]),
            (f"{chg:+.2f}%",                                                fg_today),
            (f"{pnl_pct:+.2f}%" if isinstance(pnl_pct,float) else "N/A",   fg_pnl),
            (f"A${mkt_val:,.0f}" if isinstance(mkt_val,float) else "N/A",   C["text_white"]),
            (note,                                                           C["text_dim"]),
        ]):
            c = ws.cell(row=r, column=2+i, value=v)
            sd(c, bg=row_bg, fg=fc, align="left" if i in (0,7) else "right")
        ws.row_dimensions[r].height = 18

    tot_row = 4 + len(PORTFOLIO) + 1
    ws.merge_cells(f"B{tot_row}:F{tot_row}")
    c = ws[f"B{tot_row}"]
    c.value = "TOTAL PORTFOLIO"
    c.font  = Font(name="Consolas", bold=True, color=C["gold"], size=9)
    c.fill  = hf(C["bg_mid"])
    c.alignment = Alignment(horizontal="left", vertical="center")
    tot_pnl = round((total_val-total_inv)/total_inv*100, 2) if total_inv else 0
    c2 = ws.cell(row=tot_row, column=8, value=f"A${total_val:,.0f}  ({tot_pnl:+.1f}%)")
    c2.font = Font(name="Consolas", bold=True, color=C["green"] if tot_pnl>0 else C["red"], size=9)
    c2.fill = hf(C["bg_mid"])
    c2.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[tot_row].height = 20

def build_watchlist_sheet(wb, watchlist_data):
    ws = wb.create_sheet("Watchlist")
    ws.sheet_view.showGridLines = False
    fill_bg(ws, 40, 14)
    set_widths(ws, {"A":2,"B":14,"C":12,"D":10,"E":12,"F":12,"G":14})

    ws.merge_cells("B1:G1")
    c = ws["B1"]
    c.value = "  WATCHLIST"
    c.font  = Font(name="Consolas", bold=True, color=C["accent"], size=13)
    c.fill  = hf(C["bg_dark"])
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    for i, h in enumerate(["TICKER","PRICE","TODAY %","52W HIGH","52W LOW","VS 52W HIGH"]):
        c = ws.cell(row=3, column=2+i, value=h)
        sh(c, bg=C["bg_mid"], fg=C["text_dim"], size=8)
    ws.row_dimensions[3].height = 18

    for r, ticker in enumerate(WATCHLIST, start=4):
        d     = watchlist_data.get(ticker, {})
        price = d.get("price")
        chg   = d.get("change_pct", 0)
        hi52  = d.get("52w_high")
        lo52  = d.get("52w_low")
        vs_high = round((price-hi52)/hi52*100,1) if isinstance(price,(int,float)) and isinstance(hi52,(int,float)) and hi52 else "N/A"

        row_bg = C["bg_card"] if r%2==0 else C["bg_mid"]
        fg_chg = C["green"] if chg>0 else (C["red"] if chg<0 else C["text_dim"])
        fg_vs  = C["green"] if isinstance(vs_high,float) and vs_high>-10 else (C["gold"] if isinstance(vs_high,float) and vs_high>-25 else C["red"])

        for i, (v, fc) in enumerate([
            (ticker,                                                              C["accent"]),
            (price if isinstance(price,(int,float)) else "N/A",                  C["text_white"]),
            (f"{chg:+.2f}%",                                                     fg_chg),
            (hi52 if isinstance(hi52,(int,float)) else "N/A",                    C["text_dim"]),
            (lo52 if isinstance(lo52,(int,float)) else "N/A",                    C["text_dim"]),
            (f"{vs_high:+.1f}%" if isinstance(vs_high,float) else "N/A",         fg_vs),
        ]):
            c = ws.cell(row=r, column=2+i, value=v)
            sd(c, bg=row_bg, fg=fc, align="left" if i==0 else "right")
        ws.row_dimensions[r].height = 16

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    today_str = datetime.date.today().strftime("%d %b %Y")
    print(f"\n{'─'*50}")
    print(f"  MARKET BRIEF  //  {today_str}")
    print(f"{'─'*50}")

    if ANTHROPIC_API_KEY == "YOUR_API_KEY_HERE":
        print("\nERROR: Set your ANTHROPIC_API_KEY.")
        sys.exit(1)

    print("\n[1/4] Fetching macro data...")
    macro_raw  = fetch_price_data(list(MACRO.values()))
    macro_data = {name: macro_raw.get(ticker, {}) for name, ticker in MACRO.items()}

    print("[2/4] Fetching portfolio data...")
    portfolio_data = fetch_price_data(list(PORTFOLIO.keys()))

    print("[3/4] Fetching watchlist data...")
    watchlist_data = fetch_price_data(WATCHLIST)

    audusd_val = fetch_price_data(["AUDUSD=X"]).get("AUDUSD=X", {}).get("price", 0.65)
    audusd     = audusd_val if isinstance(audusd_val,(int,float)) else 0.65

    print("[4/4] Calling Claude for analysis...")
    context = build_context(portfolio_data, macro_data, watchlist_data, audusd)
    try:
        analysis = get_claude_analysis(context)
        print("      Analysis complete")
    except Exception as e:
        print(f"      Claude error: {e}")
        analysis = {"regime": f"Claude error: {str(e)[:100]}", "sector_rotation":"N/A",
                    "portfolio_alerts":[], "top_ideas":[], "risks":"N/A", "watchlist_notes":"N/A"}

    # Always build outputs — even if Claude failed, we have live price data
    print("\nBuilding outputs...")

    # Excel
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    build_dashboard(wb, analysis, macro_data, portfolio_data, audusd, today_str)
    build_brief(wb, analysis)
    build_ideas(wb, analysis)
    build_portfolio_sheet(wb, portfolio_data, audusd)
    build_watchlist_sheet(wb, watchlist_data)
    wb.save(OUTPUT_EXCEL)
    print(f"  Excel saved: {OUTPUT_EXCEL}")

    # HTML — always written so GitHub Actions can always commit it
    html = build_html(analysis, macro_data, portfolio_data, watchlist_data, audusd, today_str)
    with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  HTML saved:  {OUTPUT_HTML}")
    print(f"\n{'─'*50}\n")

if __name__ == "__main__":
    main()
